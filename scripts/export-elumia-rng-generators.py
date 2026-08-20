"""Export RnG Bonus sheet from New Recipe.xlsx to js/elumia-rng-generators.js"""
import json
import sys
from pathlib import Path

import openpyxl

QUALITIES = ["common", "uncommon", "rare", "epic", "legendary", "max"]


def is_header_row(ws, row):
    return ws.cell(row, 1).value == "Bonus Generator"


def is_bonus_stat(stat):
    if not stat:
        return True
    text = str(stat).strip()
    return not text or text == "Bonus0" or text.startswith("Bonus")


def is_gen_start(ws, row):
    v1, v2 = ws.cell(row, 1).value, ws.cell(row, 2).value
    if v1 is None or v2 is None or v1 == "Bonus Generator":
        return False
    try:
        float(v1)
        return isinstance(v2, str)
    except (TypeError, ValueError):
        return False


def parse_sheet(ws):
    generators = []
    row = 2
    while row <= ws.max_row:
        if not is_gen_start(ws, row):
            row += 1
            continue
        gen_id = int(float(ws.cell(row, 1).value))
        gen_type = str(ws.cell(row, 2).value)
        roll_pcts = [ws.cell(row, 3 + i).value for i in range(6)]
        bonus_rows = []
        current = row
        while current <= ws.max_row:
            if current != row and is_gen_start(ws, current):
                break
            if is_header_row(ws, current):
                current += 1
                continue
            slots = []
            for slot_idx, start_col in enumerate([9, 14, 19, 24, 29]):
                stat = ws.cell(current, start_col).value
                if is_bonus_stat(stat):
                    continue
                min_val = ws.cell(current, start_col + 3).value
                max_val = ws.cell(current, start_col + 4).value
                if min_val is None and max_val is None:
                    continue
                slots.append({
                    "slot": slot_idx,
                    "quality": QUALITIES[slot_idx],
                    "stat": str(stat).strip(),
                    "group": ws.cell(current, start_col + 1).value,
                    "weight": ws.cell(current, start_col + 2).value,
                    "min": min_val,
                    "max": max_val,
                })
            if slots:
                bonus_rows.append(slots)
            current += 1
        generators.append({
            "id": gen_id,
            "type": gen_type,
            "rollPcts": roll_pcts,
            "bonusRows": bonus_rows,
        })
        row = current
    return generators


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "New Recipe.xlsx"
    out = Path(__file__).resolve().parent.parent / "js" / "elumia-rng-generators.js"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["RnG Bonus"]
    generators = parse_sheet(ws)
    out.write_text(
        "/* Auto-generated from New Recipe.xlsx — RnG Bonus sheet */\n"
        "window.ElumiaRngGenerators = "
        + json.dumps(generators, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(generators)} generators to {out}")


if __name__ == "__main__":
    main()
